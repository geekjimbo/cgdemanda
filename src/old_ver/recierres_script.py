#!/usr/bin/python

import os
import json
import subprocess
import time
import pdb
import psycopg2
import sys
import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

def log_write(New_String):
    if not(os.path.exists(logfile_dir)):
        os.mkdir(logfile_dir)
    filename = str(datetime.now().strftime('%d%m%Y')+ '.log')
    log_time = str(datetime.now().strftime('%d%m%Y %H:%M:%S') +': ')
    with open(logfile_dir + '/' + filename , 'a+') as fh1:
        fh1.write(log_time + New_String +'\n')
        time.sleep(0.1)


with open('configuracion.json', 'r') as f:
    data_json = json.load(f)
    Data_Base_dir = data_json["Variable_Database"]["source_path"] + 'xlsx/'
    postgresql_path = data_json["Variable_Database"]["postgresql_connect"]
    logfile_dir=data_json["Variable_Database"]["logfile_path"]
    ctlfile_dir=data_json["Variable_Database"]["control_path"]
log_write("Inicializacion del Scrip de Bases de Datos de Recierres de CoopeGuanacaste")
#pdb.set_trace()

ls= subprocess.Popen(['ls', '-ltr', Data_Base_dir], stdout=subprocess.PIPE,)
aws=subprocess.Popen(['awk','{print $9}'],stdin=ls.stdout,stdout=subprocess.PIPE,)
end_of_pipe = aws.stdout
print end_of_pipe
for line in end_of_pipe:
#    pdb.set_trace()
    a=str(line.strip())
    if (a== ""):
        print "nothing "
        continue
    temp_path = Data_Base_dir + a
    if (os.path.exists(temp_path)):
        print temp_path
        #pdb.set_trace()
        wb2 = load_workbook(temp_path, use_iterators=True)
#        xlsx = pd.ExcelFile(temp_path)
#        sheet1 = xlsx.parse(0)
        #print sheet1.head()
#        print sheet1.columns
        print wb2.get_sheet_names()
#        pdb.set_trace()
        sheets=wb2.get_sheet_names()
        sheet=wb2.active#get_sheet_by_name(sheets[0])
        print sheet.title
        for columns in range(2,sheet.max_column + 1):
            column_name = (str(sheet.cell(row=1, column= columns).value)).strip()
            if column_name.find('\nValue'):
                column_name = column_name.replace('\nValue', '')
            if column_name.find('_'):
                (a,b)= column_name.split('_',1)
                column_name = b
           # print temp_name
            print column_name
            time.sleep(1)
            
#        for row in sheet.iter_rows():
#            for cell in row:
#                print cell.value
#        column = sheet1.icol(0).real
#        for each row in column:
#            x=row.
#        df = pd.read_excel(xlsx, 'Sheet1')

#a = subprocess.check_output(["ls -ltr " + Data_Base_dir +" | awk '{print $9}'"])
#list=end_of_pipe[1].split('\n')
#print 'Have %d bytes in output' % len(output)
#print output
#print list[0]
#print len(list)

