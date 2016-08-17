#!/usr/bin/python

import os
import json
import subprocess
import time
import pdb
import psycopg2
import sys
import os
import pandas as pd
import pickle
from datetime import datetime
from openpyxl import load_workbook

def log_write(New_String):
    print New_String
    if not(os.path.exists(logfile_dir)):
        os.mkdir(logfile_dir)
    filename = str(datetime.now().strftime('%d%m%Y')+ '_recierres'+ '.log')
    log_time = str(datetime.now().strftime('%d%m%Y %H:%M:%S') +': ')
    with open(logfile_dir + '/' + filename , 'a+') as fh1:
        fh1.write(log_time + New_String +'\n')
        time.sleep(0.1)
#VARIABLES
my_dict=[]
my_local_dict={}

query_insert_recierres = "INSERT INTO datos_recierres (\
                name, \
                timestamp , \
                counter) VALUES (\
                %(name)s, \
                %(timestamp)s , \
                %(counter)s)"

with open('configuracion.json', 'r') as f:
    data_json = json.load(f)
    Data_Base_dir = data_json["Variable_Database"]["source_path"] + 'xlsx/'
    postgresql_path = data_json["Variable_Database"]["postgresql_connect"]
    logfile_dir=data_json["Variable_Database"]["logfile_path"]
    ctlfile_dir=data_json["Variable_Database"]["control_path"]
    procesados_dir=data_json["Variable_Database"]["processed_path"]
    xls_dir = data_json["Variable_Database"]["xls_path"]
log_write("Inicializacion del Scrip de Bases de Datos de Recierres de CoopeGuanacaste")

if not(os.path.exists(xls_dir)):
    log_write("ERROR: No se puede accesar la carpeta fuente de XLS")
    sys.exit(1)
else:
    log_write("INFO: Copiando los archivos fuente")
    commy = 'python 1recierres_conversion.py'# + xls_dir + '*.XLS ' + Data_Base_dir
    os.system(commy)
    log_write("INFO: Fin del script de conversion Archivos XLS")
#pickle.dump(my_local_dict , open(ctlfile_dir + 'recierres.txt' , 'wb'))
#pdb.set_trace()
my_local_dict=pickle.load(open(ctlfile_dir + 'recierres.txt' , 'rb'))

ls= subprocess.Popen(['ls', '-ltr', Data_Base_dir], stdout=subprocess.PIPE,)
aws=subprocess.Popen(['awk','{print $9}'],stdin=ls.stdout,stdout=subprocess.PIPE,)
end_of_pipe = aws.stdout
#print end_of_pipe

for line in end_of_pipe:
#    pdb.set_trace()
    my_item={}
    my_local_dict=pickle.load(open(ctlfile_dir + 'recierres.txt' , 'rb'))
    temp_filename=str(line.strip())    
    if (temp_filename == "" or temp_filename == "procesados" ):
        print "nothing "
        continue
    temp_path = Data_Base_dir + temp_filename
    if (os.path.exists(temp_path)):
        log_write("INFO: Cargando el Archivo: " + temp_filename + " para analisis")
        #print temp_path
        #pdb.set_trace()
        wb2 = load_workbook(temp_path, use_iterators=True)
#        xlsx = pd.ExcelFile(temp_path)
#        sheet1 = xlsx.parse(0)
        #print sheet1.head()
#        print sheet1.columns
        #print wb2.get_sheet_names()
#        pdb.set_trace()
        sheets=wb2.get_sheet_names()
        sheet=wb2.active#get_sheet_by_name(sheets[0])
        #print sheet.title
        for columns in range(2,sheet.max_column + 1):
            column_name = (str(sheet.cell(row=1, column= columns).value)).strip()
            if (column_name.find('\nValue') > 0):
                column_name = column_name.replace('\nValue', '')
            if (column_name.find('_') > 0):
                (a,b)= column_name.split('_',1)
                column_name = b
           # print temp_name
            print column_name
#            time.sleep(.3)
            for rows in range(2,sheet.max_row +1):
                row_value = (str(sheet.cell(row=rows, column= columns).value)).strip()
#                pdb.set_trace()
                if (row_value.find(' ') > 0):
                    (row_value, a)= row_value.split(' ',1)
                    log_write("INFO: Filtrando dato: " + a + " en " + temp_filename)
                if (row_value.find(',') > 0):
                    (row_value, a)= row_value.split(',',1)
                
#                    print "loca"
                #row_value = int(row_value)
                #print int(row_value)
                if not(my_local_dict.has_key(column_name)):
                    my_local_dict[column_name]= row_value
                    log_write("INFO: Agregando nueva columna a Control de recierres: " + column_name + " con un valor de contador en: " + row_value + " encontrado en el Archivo:" + temp_filename)

                    #print my_local_dict

                    temp_time = (str(sheet.cell(row=rows, column=1).value)).strip()
                    #pdb.set_trace()
                    if (temp_time.find('d') > 0):
                        row_time = datetime.strptime(temp_time, "%md%dy%Y %H:%M")
                    else:
                        row_time = datetime.strptime(temp_time, "%m/%d/%Y %H:%M")
                    my_item['timestamp'] = str(row_time)
                    my_item['name'] = column_name
                    my_item['counter'] = row_value
                    my_dict.append(my_item)
                    my_item={}
                    print "my_dict: " ,  my_dict
                    #pdb.set_trace()
                else:
                    if int(my_local_dict[column_name]) < int(row_value):
                        log_write("INFO: Cambiando contador de " + column_name + " de " + my_local_dict[column_name] + " a " + row_value )
                        #pdb.set_trace()
                        my_local_dict[column_name]= row_value
                        #print my_local_dict
                        temp_time = (str(sheet.cell(row=rows, column=1).value)).strip()
                        #pdb.set_trace()
                        if (temp_time.find('d') > 0):
                            row_time = datetime.strptime(temp_time, "%md%dy%Y %H:%M")
                        else:
                            row_time = datetime.strptime(temp_time, "%m/%d/%Y %H:%M")
                        my_item['timestamp'] = str(row_time)
                        my_item['name'] = column_name
                        my_item['counter'] = row_value
                        my_dict.append(my_item)
                        my_item={}
                        #print  "my_dict: " , my_dict
                       # pdb.set_trace()
    log_write("INFO: Cambiando de Archivo")
    #print('aca cambiamos de archivo')
    print my_dict
    subprocess.call(["cp", (ctlfile_dir + 'recierres.txt') ,(ctlfile_dir + 'recierres.txt.old')])
    pickle.dump(my_local_dict , open(ctlfile_dir + 'recierres.txt' , 'wb'))
    my_local_dict= {}
    log_write("INFO: Escribiendo archivo interno de Control de reccieress. Cantidad de lineas: " + str(len(my_dict)))
    con = psycopg2.connect(postgresql_path)
    cur = con.cursor()
    cur.executemany(query_insert_recierres, my_dict)
    con.commit()
    log_write("Cerrando Conexion con DB para "+ temp_filename)
    con.close()
    my_dict=[]
    subprocess.call(["mv", temp_path , procesados_dir])
    log_write("Moviendo archivo "+ temp_filename +" hacia: " + procesados_dir)
#    log_write("Escribiendo en Bases de Datos la informacion de " + i["name"] +". Cantidad de lineas: " + str(len(my_dict)))

    #aca se guarda en DB

#pickle.dump(my_local_dict , open(ctlfile_dir + 'recierres.txt' , 'wb'))
#        for row in sheet.iter_rows():
#            for cell in row:
#                print cell.value
#        column = sheet1.icol(0).real
#        for each row in column:
#            x=row.
#        df = pd.read_excel(xlsx, 'Sheet1')

#a = subprocess.check_output(["ls -ltr " + Data_Base_dir +" | awk '{print $9}'"])
#list=end_of_pipe[1].split('\n')
#print 'Have %d bytes in output' % len(output)
#print output
#print list[0]
#print len(list)

